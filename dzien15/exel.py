import openpyxl
from pprint import pprint


plik = openpyxl.load_workbook("example2.xlsx")

# print(plik)

arkusze = plik.sheetnames
print(arkusze)

aktywny_arkusz = plik.get_sheet_by_name("Owoce")
print(f"Aktywny arkusz: {aktywny_arkusz.title}")

# kokórki

komorka = aktywny_arkusz["A1"]
print(komorka)
print(komorka.value)

# from openpyxl.cell.cell import Cell
# assert isinstance(komorka, Cell.)

# wspolrzedne
print(f"Adres komorki: {komorka.coordinate}")
print(f"Kolumna komorki: {komorka.column}")
print(f"Wiersz komorki {komorka.row}")

print(aktywny_arkusz.cell(row=2, column=2))

# zmiana liter kolumn na liczbe i vive wersa
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(123))
print(column_index_from_string("DS"))

# (wspolrzedne) dolnego prawego rogu arkusza dokad sa dane
print("Ostatnia koluna to:")
print(aktywny_arkusz.max_column)
print(get_column_letter(aktywny_arkusz.max_column))
print("Ostatni wiersz to;")
print(aktywny_arkusz.max_row)

pprint(aktywny_arkusz["A1":"C7"], indent=4)

for wiersz in aktywny_arkusz["A1":"C9"]:
    for kom in wiersz:
        print(kom.value, end='  ')
    print()

aktywny_arkusz["B9"] = "wpis z Pythona"
plik.save("example2.xlsx")
plik.close()